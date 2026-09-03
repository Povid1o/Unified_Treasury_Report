"""Сборка Excel-книги отчёта «Ставки ОФЗ».

Формирует готовый `excel_ofz/dist/ОФЗ_ставки.xlsm` — с листами, именованными
диапазонами, справочником индексов CBonds, вшитым модулем VBA из
`ofz_report.bas` и кнопками запуска. Нужен только Python: ни Excel, ни Windows
для сборки не требуется.

Справочник индексов, список сроков, глубина истории и перечень неполучаемых
показателей импортируются из рабочего Python-кода, чтобы Excel-версия не
разъезжалась с `reports/ofz_rates/etl.py`.

Запуск:
    python excel_ofz/build_workbook.py
"""
from __future__ import annotations

import sys
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BASE_DIR))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import quote_sheetname  # noqa: E402
from openpyxl.workbook.defined_name import DefinedName  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402

from CBonds_API.cbonds_api_test import AVAILABLE_INDICES  # noqa: E402
from reports.ofz_rates.etl import (  # noqa: E402
    BI_COLUMNS,
    KNOWN_GAPS,
    LOOKBACK_DAYS,
    OFZ_YIELD_TENORS,
)

sys.path.insert(0, str(Path(__file__).resolve().parent))
from vba_container import (  # noqa: E402
    DOC_BASE_WORKBOOK,
    DOC_BASE_WORKSHEET,
    MODULE_DOCUMENT,
    VbaModule,
    build_vba_project,
)
from xlsm_package import Button, to_xlsm  # noqa: E402
from check_vba import check as check_vba  # noqa: E402

HERE = Path(__file__).resolve().parent
DIST_DIR = HERE / "dist"
OUTPUT_XLSM = DIST_DIR / "ОФЗ_ставки.xlsm"
BAS_SOURCE = HERE / "ofz_report.bas"
BAS_CP1251 = DIST_DIR / "ofz_report_cp1251.bas"

MODULE_NAME = "modOFZ"
PROJECT_NAME = "VBAProject"
CODE_PAGE = 1251

BUTTONS = [
    Button("Сформировать отчёт", "RunOfzReport", 180, "1F4E79"),
    Button("Выгрузить XLSX", "ExportOfzXlsx", 150, "2E6DA4"),
    Button("Выгрузить CSV", "ExportOfzCsv", 140, "4A7EA8"),
    Button("Проверить связь", "TestConnection", 150, "7F7F7F"),
]

SH_PARAMS = "Параметры"
SH_REPORT = "Отчёт"
SH_LOG = "Лог"
SH_INDEX = "Индексы"
SH_GAPS = "Пропуски"

ACCENT = "1F4E79"
HEAD_FILL = PatternFill("solid", fgColor=ACCENT)
SECTION_FILL = PatternFill("solid", fgColor="DCE6F1")
INPUT_FILL = PatternFill("solid", fgColor="FFF9E0")
HEAD_FONT = Font(bold=True, color="FFFFFF")
SECTION_FONT = Font(bold=True, color=ACCENT)
HINT_FONT = Font(italic=True, size=9, color="808080")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _input_cell(ws, ref: str, value=None, number_format: str = "@"):
    cell = ws[ref]
    if value is not None:
        cell.value = value
    cell.fill = INPUT_FILL
    cell.border = BOX
    cell.number_format = number_format
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return cell


def _section(ws, row: int, title: str, width: int = 3):
    ws.cell(row=row, column=1, value=title).font = SECTION_FONT
    for col in range(1, width + 1):
        ws.cell(row=row, column=col).fill = SECTION_FILL


def _hint(ws, row: int, col: int, text: str):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = HINT_FONT
    cell.alignment = Alignment(vertical="center")


def _header_row(ws, headers, row: int = 1):
    for i, name in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=name)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def build_params_sheet(ws) -> dict[str, str]:
    """Лист с настройками запуска. Возвращает карту «имя диапазона -> ссылка»."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 78

    ws["A1"] = "Ставки ОФЗ — выгрузка из CBonds API"
    ws["A1"].font = Font(bold=True, size=16, color=ACCENT)
    _hint(ws, 2, 1, "Заполните параметры и нажмите «Сформировать отчёт». Результат — на листе «Отчёт», ход работы — на листе «Лог».")

    _section(ws, 4, "Доступ к CBonds")
    ws["A5"] = "Логин"
    _input_cell(ws, "B5")
    _hint(ws, 5, 3, "Значение CBONDS_LOGIN")
    ws["A6"] = "Пароль"
    _input_cell(ws, "B6")
    _hint(ws, 6, 3, "Можно оставить пустым — тогда пароль спросят при запуске и он не сохранится в файле (так безопаснее).")

    _section(ws, 8, "Период")
    ws["A9"] = "Режим"
    mode = _input_cell(ws, "B9", "1", number_format="General")
    mode.alignment = Alignment(horizontal="center", vertical="center")
    _hint(ws, 9, 3, "1 — одна дата + история назад   |   2 — интервал дат   |   3 — список конкретных дат")

    ws["A10"] = "Дата отчёта"
    _input_cell(ws, "B10")
    _hint(ws, 10, 3, "Режим 1. YYYY-MM-DD. Пусто = сегодня.")

    ws["A11"] = "Глубина истории, дней"
    _input_cell(ws, "B11", LOOKBACK_DAYS, number_format="0")
    _hint(ws, 11, 3, f"Режим 1. Сколько календарных дней назад от даты отчёта (по умолчанию {LOOKBACK_DAYS}).")

    ws["A12"] = "Начало интервала"
    _input_cell(ws, "B12")
    _hint(ws, 12, 3, "Режим 2. YYYY-MM-DD.")

    ws["A13"] = "Конец интервала"
    _input_cell(ws, "B13")
    _hint(ws, 13, 3, "Режим 2. YYYY-MM-DD.")

    ws["A14"] = "Список дат"
    _input_cell(ws, "B14")
    _hint(ws, 14, 3, "Режим 3. Через запятую: 2026-06-01,2026-06-15,2026-07-01 — даты не обязаны идти подряд.")

    _section(ws, 16, "Показатели")
    ws["A17"] = "Сроки кривой ОФЗ"
    _input_cell(ws, "B17", ",".join(OFZ_YIELD_TENORS))
    _hint(ws, 17, 3, "Через запятую. Допустимые сроки — колонка A листа «Индексы» (префикс RUB_Yield_Curve_).")

    ws["A18"] = "Лимит записей на срок"
    _input_cell(ws, "B18", 200, number_format="0")
    _hint(ws, 18, 3, "Сколько значений максимум забирать по одному сроку за запрос.")

    _section(ws, 20, "Выгрузка для BI")
    ws["A21"] = "Путь для XLSX"
    _input_cell(ws, "B21")
    _hint(ws, 21, 3, "Кнопка «Выгрузить XLSX»: отдельный файл с одним листом-выгрузкой. Пусто = ofz_report.xlsx рядом с книгой.")

    ws["A22"] = "Путь для CSV"
    _input_cell(ws, "B22")
    _hint(ws, 22, 3, "Кнопка «Выгрузить CSV»: UTF-8 с BOM, байт в байт как у Python-версии. Пусто = ofz_report.csv рядом с книгой.")

    _section(ws, 24, "Статус")
    ws["A25"] = "Последний запуск"
    status = ws["B25"]
    status.value = "Отчёт ещё не запускался"
    status.font = Font(bold=True)
    status.alignment = Alignment(vertical="center")

    dv_mode = DataValidation(type="list", formula1='"1,2,3"', allow_blank=False)
    ws.add_data_validation(dv_mode)
    dv_mode.add(ws["B9"])

    for row in range(5, 26):
        ws.row_dimensions[row].height = 18
    # Строки под кнопки: высота с запасом, чтобы фигуры не наезжали друг на друга.
    for row in range(28, 28 + len(BUTTONS)):
        ws.row_dimensions[row].height = 26

    q = quote_sheetname(ws.title)
    return {
        "p_login": f"{q}!$B$5",
        "p_password": f"{q}!$B$6",
        "p_mode": f"{q}!$B$9",
        "p_date": f"{q}!$B$10",
        "p_lookback": f"{q}!$B$11",
        "p_date_from": f"{q}!$B$12",
        "p_date_to": f"{q}!$B$13",
        "p_dates": f"{q}!$B$14",
        "p_tenors": f"{q}!$B$17",
        "p_limit": f"{q}!$B$18",
        "p_xlsx_path": f"{q}!$B$21",
        "p_csv_path": f"{q}!$B$22",
        "p_status": f"{q}!$B$25",
    }


def build_report_sheet(ws):
    """Лист результата — сырая выгрузка под загрузку в BI, без оформления.

    Ровно то, что писал `save_report()` в CSV: строка заголовка + строки данных,
    те же колонки в том же порядке. Никаких заливок, объединений, итогов,
    закреплённых областей и подписей — иначе парсер BI споткнётся о лишнее.
    """
    for i, name in enumerate(BI_COLUMNS, start=1):
        ws.cell(row=1, column=i, value=name)

    widths = {"A": 14, "B": 20, "C": 12, "D": 8, "E": 12, "F": 12}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Единственный формат на листе, и он не косметический: без него Excel
    # распознает "2026-07-01" как дату и подставит локальный формат вместо ISO.
    # fvalue намеренно оставлен General — в ячейке лежит то же число, что ушло
    # бы в CSV (13.5, а не "13.50").
    ws.column_dimensions["A"].number_format = "@"
    ws["A1"].number_format = "@"


def build_log_sheet(ws):
    _header_row(ws, ["Время", "Уровень", "Сообщение"])
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 150


def build_index_sheet(ws):
    """Справочник type_id CBonds — VBA берёт ID отсюда, а не из зашитых констант."""
    _header_row(ws, ["Ключ индекса", "type_id"])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
    for i, (key, type_id) in enumerate(sorted(AVAILABLE_INDICES.items()), start=2):
        ws.cell(row=i, column=1, value=key)
        ws.cell(row=i, column=2, value=type_id).number_format = "0"


def build_gaps_sheet(ws):
    """Показатели, которых нет ни в одном доступном аккаунту эндпоинте."""
    _header_row(ws, ["name_group", "name_st", "Причина"])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 150
    for i, gap in enumerate(KNOWN_GAPS, start=2):
        ws.cell(row=i, column=1, value=gap["name_group"])
        ws.cell(row=i, column=2, value=gap["name_st"])
        cell = ws.cell(row=i, column=3, value=" ".join(gap["reason"].split()))
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[i].height = 60


def convert_bas_to_cp1251() -> Path | None:
    """Запасной путь: подключить модуль вручную, если что-то пойдёт не так.

    VBE импортирует .bas в системной кодировке, для русской Windows это cp1251.
    """
    if not BAS_SOURCE.exists():
        return None
    BAS_CP1251.write_bytes(BAS_SOURCE.read_text(encoding="utf-8").encode(f"cp{CODE_PAGE}", "replace"))
    return BAS_CP1251


def build_vba(sheet_count: int) -> bytes:
    """Проект VBA: модули документов под каждый лист + модуль отчёта.

    Модули документов пустые, но нужны: Excel сопоставляет их листам по
    codeName, и без них проект не соответствует книге.
    """
    modules = [VbaModule("ThisWorkbook", "", MODULE_DOCUMENT, DOC_BASE_WORKBOOK)]
    for i in range(1, sheet_count + 1):
        modules.append(VbaModule(f"Sheet{i}", "", MODULE_DOCUMENT, DOC_BASE_WORKSHEET))
    modules.append(VbaModule(MODULE_NAME, BAS_SOURCE.read_text(encoding="utf-8")))
    return build_vba_project(PROJECT_NAME, modules, CODE_PAGE)


def main() -> Path:
    # Компилятора VBA вне Windows нет, поэтому синтаксис проверяем сами:
    # ошибка вида «Only comments may appear after End Sub» иначе всплывёт
    # только при открытии готовой книги в Excel.
    problems = check_vba(BAS_SOURCE)
    if problems:
        print(f"{BAS_SOURCE.name}: замечаний — {len(problems)}, сборка остановлена")
        for problem in problems:
            print(f"  строка {problem.line}: {problem.message}")
            print(f"    {problem.text[:110]}")
        raise SystemExit(1)

    DIST_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_params = wb.active
    ws_params.title = SH_PARAMS
    names = build_params_sheet(ws_params)

    build_report_sheet(wb.create_sheet(SH_REPORT))
    build_log_sheet(wb.create_sheet(SH_LOG))
    build_index_sheet(wb.create_sheet(SH_INDEX))
    build_gaps_sheet(wb.create_sheet(SH_GAPS))

    for name, ref in names.items():
        wb.defined_names.add(DefinedName(name, attr_text=ref))

    # codeName связывает лист с его модулем в проекте VBA.
    wb.code_name = "ThisWorkbook"
    for i, ws in enumerate(wb.worksheets, start=1):
        ws.sheet_properties.codeName = f"Sheet{i}"

    wb.active = 0
    staging = DIST_DIR / "_staging.xlsx"
    wb.save(staging)

    to_xlsm(staging, OUTPUT_XLSM, build_vba(len(wb.worksheets)), BUTTONS, first_button_row=27)
    staging.unlink()

    bas = convert_bas_to_cp1251()
    print(f"Книга с макросами: {OUTPUT_XLSM}  ({OUTPUT_XLSM.stat().st_size // 1024} КБ)")
    print(f"Модуль VBA (запасной, для ручного импорта): {bas}")
    print(f"Кнопок: {len(BUTTONS)} | индексов в справочнике: {len(AVAILABLE_INDICES)} | "
          f"сроков по умолчанию: {len(OFZ_YIELD_TENORS)}")
    return OUTPUT_XLSM


if __name__ == "__main__":
    main()
