"""Чтение Excel-файлов, устойчивое к артефактам выгрузок из OLAP-кубов.

Зачем: значительная часть входных файлов — наполовину выгрузки из OLAP-кубов,
и обычный pd.read_excel на них возвращает пустоту вместо данных. Причин две,
обе воспроизводимы:

1. ФОРМУЛЫ БЕЗ КЭША. Ячейки содержат формулы (в выгрузках из кубов — обычно
   CUBEVALUE/CUBEMEMBER). pandas читает Excel через openpyxl с data_only=True,
   то есть берёт не саму формулу, а ЗАКЭШИРОВАННЫЙ результат её вычисления,
   сохранённый в файле. Если кэша нет (файл выгружен без подключения к кубу,
   либо его пересохранил инструмент, который кэш не пишет — например любой
   скрипт на openpyxl), pandas возвращает NaN. Данных в файле физически нет:
   CUBEVALUE без живого подключения к кубу пересчитать невозможно, поэтому
   восстановить их кодом нельзя — можно только внятно об этом сказать
   (см. describe_uncached_formulas), а не выдавать молча пустой отчёт.

2. ОБЪЕДИНЁННЫЕ ЯЧЕЙКИ. В объединённом блоке значение хранится только в
   левой верхней ячейке, остальные — None. Маркеры и названия статей в
   выгрузках часто оказываются в таких блоках, и поиск по колонке A их не
   находит. Это чинится: read_matrix заполняет весь объединённый диапазон
   значением левой верхней ячейки (ровно то, что показывает Excel).
"""
from pathlib import Path
from typing import Any, Iterator, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


class ExcelSourceError(RuntimeError):
    """Входной Excel-файл невозможно прочитать осмысленно."""


# Функции OLAP-куба: если такая формула осталась без кэша, значение
# принципиально невосстановимо без подключения к кубу.
CUBE_FUNCTIONS = ("CUBEVALUE", "CUBEMEMBER", "CUBESET", "CUBERANKEDMEMBER", "CUBEKPIMEMBER")


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def find_uncached_formulas(
    file_path: Path, sheet_name: Any = 0, limit: int = 20
) -> Tuple[int, List[str], bool]:
    """Ищет ячейки с формулой, но без закэшированного значения.

    Возвращает (сколько_всего, примеры_адресов, есть_ли_среди_них_CUBE_функции).
    Именно эти ячейки pandas отдаёт как NaN — см. модульную docstring.
    """
    wb_formulas = load_workbook(file_path, data_only=False)
    wb_values = load_workbook(file_path, data_only=True)
    try:
        ws_f = wb_formulas[sheet_name] if isinstance(sheet_name, str) else wb_formulas.worksheets[sheet_name]
        ws_v = wb_values[sheet_name] if isinstance(sheet_name, str) else wb_values.worksheets[sheet_name]

        total = 0
        examples: List[str] = []
        has_cube = False
        for row_f, row_v in zip(ws_f.iter_rows(), ws_v.iter_rows()):
            for cell_f, cell_v in zip(row_f, row_v):
                if _is_formula(cell_f.value) and cell_v.value is None:
                    total += 1
                    if len(examples) < limit:
                        examples.append(f"{cell_f.coordinate}: {cell_f.value[:60]}")
                    if any(fn in cell_f.value.upper() for fn in CUBE_FUNCTIONS):
                        has_cube = True
        return total, examples, has_cube
    finally:
        wb_formulas.close()
        wb_values.close()


def describe_uncached_formulas(file_path: Path, sheet_name: Any = 0) -> Optional[str]:
    """Человекочитаемое объяснение, почему файл прочитался пустым, либо None."""
    total, examples, has_cube = find_uncached_formulas(file_path, sheet_name)
    if total == 0:
        return None

    lines = [
        f"В файле {Path(file_path).name} найдено {total} ячеек с формулами, "
        "у которых не сохранено вычисленное значение — именно поэтому они "
        "читаются как пустые.",
    ]
    if has_cube:
        lines.append(
            "Среди них есть функции OLAP-куба (CUBEVALUE/CUBEMEMBER). Их значения "
            "невозможно восстановить из файла — они вычисляются только при живом "
            "подключении к кубу."
        )
    lines.append(
        "Что делать: открыть файл в Excel с подключением к кубу, обновить данные "
        "(Данные → Обновить всё) и пересохранить; либо выделить диапазон и "
        "вставить его как значения (Специальная вставка → Значения), после чего "
        "сохранить файл заново."
    )
    lines.append("Примеры ячеек: " + "; ".join(examples))
    return "\n".join(lines)


def _worksheet_to_frame(ws) -> pd.DataFrame:
    """Лист -> DataFrame без заголовков, с развёрнутыми объединёнными ячейками:
    весь объединённый диапазон получает значение левой верхней ячейки, как это
    и выглядит в Excel (openpyxl отдаёт его только в левой верхней)."""
    data = [list(row) for row in ws.iter_rows(values_only=True)]
    if not data:
        return pd.DataFrame()

    df = pd.DataFrame(data)
    for rng in ws.merged_cells.ranges:
        top_left = df.iat[rng.min_row - 1, rng.min_col - 1]
        if top_left is None:
            continue
        for r in range(rng.min_row - 1, rng.max_row):
            for c in range(rng.min_col - 1, rng.max_col):
                df.iat[r, c] = top_left
    return df


def iter_sheet_matrices(file_path: Path) -> Iterator[Tuple[str, pd.DataFrame]]:
    """Проходит ПО ВСЕМ листам книги, отдавая (имя_листа, матрица значений).

    Нужно отчётам, которые не знают заранее, на каком листе лежат данные:
    выгрузки из кубов кладут перед нужным листом служебные (кэш сводной,
    параметры подключения, титульный), и чтение "первого листа по умолчанию"
    молча возвращает пустоту. Книга загружается один раз на все листы.
    """
    wb = load_workbook(Path(file_path), data_only=True)
    try:
        for name in wb.sheetnames:
            yield name, _worksheet_to_frame(wb[name])
    finally:
        wb.close()


def read_matrix(file_path: Path, sheet_name: Any = 0, logger: Any = None) -> pd.DataFrame:
    """Читает лист как голую матрицу (аналог pd.read_excel(header=None)),
    но разворачивает объединённые ячейки: весь объединённый диапазон
    получает значение левой верхней ячейки — так же, как это выглядит в
    Excel. Без этого маркеры и названия статей, попавшие в объединённые
    блоки, не находятся поиском по колонке (см. модульную docstring).

    Если переданы logger и в файле есть формулы без кэша — предупреждает:
    такие ячейки читаются как пустые, то есть часть данных потеряна молча.
    Если лист оказался полностью пустым — поднимает ExcelSourceError с
    объяснением причины.
    """
    file_path = Path(file_path)
    wb = load_workbook(file_path, data_only=True)
    try:
        try:
            ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[sheet_name]
        except (KeyError, IndexError) as exc:
            raise ExcelSourceError(
                f"Лист {sheet_name!r} не найден в {file_path.name}. "
                f"Листы в файле: {wb.sheetnames}. "
                "Если выгрузка называет лист иначе — поправьте SHEET_NAME в ETL этого отчёта."
            ) from exc

        df = _worksheet_to_frame(ws)
    finally:
        wb.close()

    if df.empty:
        raise ExcelSourceError(f"Лист {sheet_name!r} в файле {file_path.name} пуст.")

    if df.isna().all().all():
        explanation = describe_uncached_formulas(file_path, sheet_name)
        if explanation:
            raise ExcelSourceError(explanation)
        raise ExcelSourceError(
            f"Лист {sheet_name!r} в файле {file_path.name} не содержит данных."
        )

    if logger is not None:
        total, _examples, has_cube = find_uncached_formulas(file_path, sheet_name)
        if total:
            logger.warning(
                "В файле %s: %d ячеек с формулами без сохранённого значения%s — "
                "они прочитаются как пустые. Обновите и пересохраните файл в Excel, "
                "либо вставьте данные как значения.",
                file_path.name, total,
                " (в том числе функции OLAP-куба)" if has_cube else "",
            )

    return df
