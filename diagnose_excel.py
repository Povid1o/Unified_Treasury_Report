"""Диагностика входного Excel-файла: почему он читается пустым.

Запуск:
    python diagnose_excel.py "путь/к/файлу.xlsx"

Без аргумента спросит путь (можно перетащить файл в окно терминала).

Проверяет по каждому листу три вещи, из-за которых выгрузки из OLAP-кубов
читаются как пустота (подробности — в common/excel_io.py):
  1. формулы без сохранённого значения (CUBEVALUE и т.п.) — pandas отдаёт NaN;
  2. объединённые ячейки — значение только в левой верхней, остальные пустые;
  3. лист, где данных нет вовсе.
"""
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit(
        "Не установлен openpyxl. Запустите скрипт интерпретатором из .venv "
        "(см. инструкцию: .venv/bin/python или .venv\\Scripts\\python.exe)."
    )

CUBE_FUNCTIONS = ("CUBEVALUE", "CUBEMEMBER", "CUBESET", "CUBERANKEDMEMBER", "CUBEKPIMEMBER")


def _clean_path(raw: str) -> Path:
    """Убирает кавычки и экранирование пробелов, которые добавляет терминал
    при перетаскивании файла в окно."""
    raw = raw.strip()
    for quote in ('"', "'"):
        if len(raw) > 1 and raw.startswith(quote) and raw.endswith(quote):
            raw = raw[1:-1]
    return Path(raw.replace("\\ ", " "))


def diagnose(path: Path) -> None:
    print(f"\nФайл: {path}")
    size_mb = path.stat().st_size / 1024 / 1024
    print(f"Размер: {size_mb:.2f} МБ")

    try:
        wb_values = load_workbook(path, data_only=True)
        wb_formulas = load_workbook(path, data_only=False)
    except Exception as exc:
        print(f"\nНе удалось открыть файл через openpyxl: {exc}")
        print(
            "Если файл в старом формате .xls — пересохраните его в Excel как .xlsx.\n"
            "Если он защищён паролем — снимите защиту."
        )
        return

    print(f"Листов: {len(wb_values.sheetnames)} — {wb_values.sheetnames}\n")

    verdicts = []
    filled_by_sheet = {}
    for name in wb_values.sheetnames:
        ws_v, ws_f = wb_values[name], wb_formulas[name]

        filled = uncached = cached = 0
        has_cube = False
        examples = []

        for row_v, row_f in zip(ws_v.iter_rows(), ws_f.iter_rows()):
            for cell_v, cell_f in zip(row_v, row_f):
                is_formula = isinstance(cell_f.value, str) and cell_f.value.startswith("=")
                if cell_v.value is not None:
                    filled += 1
                    if is_formula:
                        cached += 1
                elif is_formula:
                    uncached += 1
                    if any(fn in cell_f.value.upper() for fn in CUBE_FUNCTIONS):
                        has_cube = True
                    if len(examples) < 5:
                        examples.append(f"{cell_f.coordinate}: {cell_f.value[:70]}")

        merged = len(ws_f.merged_cells.ranges)
        filled_by_sheet[name] = filled

        print(f"── Лист «{name}» ({ws_v.max_row} строк × {ws_v.max_column} колонок)")
        print(f"   Ячеек со значением, которое увидит pandas: {filled}")
        if cached:
            print(f"   Из них формул с сохранённым значением: {cached} (читаются нормально)")
        print(f"   Формул БЕЗ сохранённого значения: {uncached}" + (" ← ПРИЧИНА ПУСТОТЫ" if uncached else ""))
        if has_cube:
            print("   Среди них функции OLAP-куба (CUBEVALUE/CUBEMEMBER)")
        print(f"   Объединённых диапазонов: {merged}" + (" ← могут прятать маркеры" if merged else ""))
        for ex in examples:
            print(f"      {ex}")

        if filled == 0 and uncached > 0:
            verdicts.append((name, "формулы без кэша — данных в файле нет"))
        elif uncached > 0:
            verdicts.append((name, f"частичная потеря: {uncached} ячеек пустые"))
        elif filled == 0:
            verdicts.append((name, "лист пуст"))
        print()

    wb_values.close()
    wb_formulas.close()

    print("═" * 60)

    # Самая частая причина: данные лежат не на первом листе, а отчёт читает
    # лист по умолчанию (первый). Проверяем это отдельно и явно.
    sheets = list(filled_by_sheet)
    first_sheet = sheets[0]
    richest = max(sheets, key=lambda s: filled_by_sheet[s])
    if len(sheets) > 1 and richest != first_sheet and filled_by_sheet[first_sheet] < filled_by_sheet[richest] / 10:
        print("ВНИМАНИЕ: данные лежат НЕ на первом листе.")
        print(f"  Первый лист «{first_sheet}»: заполненных ячеек {filled_by_sheet[first_sheet]}")
        print(f"  Больше всего данных на «{richest}»: {filled_by_sheet[richest]}")
        print(
            "  Отчёт, который читает лист по умолчанию, получит пустоту.\n"
            "  Для NIM это уже исправлено — он ищет лист по маркерам.\n"
            "  Для ЧПД лист задан в reports/chpd/etl.py (SHEET_NAME),\n"
            "  для BalanceStruct — в reports/balance_struct/etl.py (SHEET_NAME):\n"
            f"  сверьте эти значения со списком листов выше.\n"
        )

    if not verdicts:
        print("Формул без кэша нет — значения в файле сохранены.")
        if len(sheets) > 1:
            print("Если отчёт всё равно выходит пустым, причина скорее всего в выборе листа (см. выше).")
        else:
            print("Причина пустоты другая — покажите этот вывод и сам файл.")
        return

    print("ИТОГ:")
    for name, verdict in verdicts:
        print(f"  • «{name}»: {verdict}")
    print(
        "\nЕсли причина — формулы без кэша, файл нужно починить на стороне источника:\n"
        "  1) открыть в Excel с подключением к кубу;\n"
        "  2) Данные → Обновить всё;\n"
        "  3) сохранить.\n"
        "Либо выделить диапазон → копировать → Специальная вставка → Значения → сохранить.\n"
        "Пересчитать эти значения из файла невозможно: их там физически нет."
    )


def main() -> None:
    if len(sys.argv) > 1:
        raw = " ".join(sys.argv[1:])
    else:
        raw = input("Путь к Excel-файлу (можно перетащить файл сюда): ")

    path = _clean_path(raw)
    if not path.exists():
        sys.exit(f"Файл не найден: {path}")
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        print(f"Внимание: расширение {path.suffix!r} — openpyxl читает только .xlsx/.xlsm.")

    diagnose(path)


if __name__ == "__main__":
    main()
